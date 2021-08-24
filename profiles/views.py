from django.shortcuts import redirect, render
from django.http import HttpResponseRedirect
from django.views.generic.list import ListView
from allauth.account.models import EmailAddress,EmailConfirmation
from django.views.generic import TemplateView, UpdateView
from django.http import JsonResponse
from allauth.account.adapter import get_adapter
from .models import User
from rice.models import Files
from django.contrib import messages
from rice.utils import get_collection_handle, get_db_handle
from django.conf import settings
import openpyxl
import time
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.mixins import UserPassesTestMixin

def staff_check(user):
    return user.is_active and user.is_staff

class TestMixin1(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_active and self.request.user.is_staff

class AdminPanelUserView(TestMixin1, ListView):
    context_object_name = 'users'
    queryset = EmailAddress.objects.filter(primary = True).order_by('user')
    template_name = 'profiles/users_admin.html'

    def post(self, request, *args, **kwargs):
        selected = request.POST.getlist('id') or None
        action = request.POST.get('action') or None
        if(not selected):
            messages.warning(request, 'Items must be selected in order to perform actions on them. No items have been changed.')
        elif(not action):
            messages.warning(request, 'No action selected.')
        else:
            if action == 'approval_user':
                for item in selected:
                    pk = int(item)
                    email_address = EmailAddress.objects.get(user=pk)
                    email_address.send_confirmation(request)
                    messages.success(request, 'User details updated.')
        return HttpResponseRedirect(request.get_full_path())

class AdminPanelFileManageView(TestMixin1, ListView):
    context_object_name = 'files'
    queryset = Files.objects.all()
    template_name = 'profiles/file_manage_admin.html'

class AdminPanelFilesView(TestMixin1, TemplateView):
    template_name = 'profiles/files_admin.html'

@user_passes_test(staff_check)
def rice_upload(request):
    t0 = time.time()
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        type_file = request.POST['typeFile']
        type_file_first = {
            'BI': settings.FIRST_ROW_BIO,
            'CH': settings.FIRST_ROW_CHEM,
            'SP': settings.FIRST_ROW_SAMPLE,
            'NU': settings.FIRST_ROW_NUTRI,
            'PH': settings.FIRST_ROW_PHY,
        }

        if(type_file == 'MD'):
            db, _ = get_db_handle(settings.DB_NAME, settings.HOST_MONGODB )
            collection = get_collection_handle(db, settings.COLLECTION_NAME_META)
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            rows = worksheet.rows
            first_row = [cell.value for cell in next(rows)]
            for row in rows:
                record = {}
                for key, cell in zip(first_row, row):
                    record[key] = cell.value
                    collection.update({"fieldID":record["fieldID"]}, {"$set": record}, upsert = True)
            return JsonResponse({'ex': True})
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.active
        rows = worksheet.rows
        first_row = [cell.value for cell in next(rows)]
        if(type_file_first[type_file] != first_row):
            return JsonResponse({'ex': False})
        else:
            obj = Files.objects.create(type_file=type_file, user=request.user,  upload=excel_file)
            a = obj.TYPE_FILE

            for row in a:
                if obj.type_file == row[0]:
                    coll = row[1]


            db, _ = get_db_handle(settings.DB_NAME, settings.HOST_MONGODB )
            collection = get_collection_handle(db, settings.COLLECTION_NAME)
            wb = openpyxl.load_workbook(obj.upload.path, read_only=True)
            worksheet = wb.active
            rows = worksheet.rows
            first_row = [cell.value for cell in next(rows)]
            nutri = collection.find({ coll: { '$exists': True } })
            nutri_list = list(nutri)
            for row in rows:
                record = {}
                for key, cell in zip(first_row, row):
                    if cell.data_type == 's':
                        record[key] = cell.value.strip()
                    elif isinstance(cell.value, float):
                        record[key] = float(cell.value)
                    else:
                        record[key] = cell.value
                if coll == 'Sampleinfo':
                    collection.update({'cropSampleID': record["cropSampleID"]},  {'$set': {'cropSampleID': record["cropSampleID"], coll: record}}, upsert = True)

                else:
                    # nutri = collection.find({'cropSampleID': record["cropSampleID"], f"{coll}.nutritionDBID": record['nutritionDBID']})
                    # nutri_list = list(nutri)

                    if nutri_list == []:
                        collection.update({'cropSampleID': record["cropSampleID"]},{'$push':{coll:record}}, upsert = True)
                    else:
                    # collection.update_one ({'cropSampleID': record["cropSampleID"]},{'$set':{f"{coll}.$[elem]":record}}, array_filters = [ { "elem.nutritionDBID": record['nutritionDBID'] } ])
                        collection.find_one_and_update({'cropSampleID': record["cropSampleID"]},{'$set':{f"{coll}.$[elem]":record}}, array_filters = [ { "elem.nutritionDBID": record['nutritionDBID'] } ], upsert = True)
    t1 = time.time()
    print(t1-t0)
    return JsonResponse({'ex': True })


